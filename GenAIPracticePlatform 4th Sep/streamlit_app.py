"""GenAI Lab - interactive practice platform (Streamlit Community Cloud ready).

Per task: prompt + progressive hints + code editor + reveal-solution + auto-scored
Check + a REAL "Run" that executes the learner's code in this app's Python
interpreter (see README "Where learner code runs").

Content is data: add folders under ./content/<lab>/ each with a practice.yaml.
"""
from __future__ import annotations

import contextlib
import importlib.util
import io
import json
import os
import re
import sys
import tempfile
import time
import traceback
from pathlib import Path

import streamlit as st

try:
    import yaml
except ModuleNotFoundError:
    st.set_page_config(page_title="Lab Practice", page_icon="🧪")
    st.error("**PyYAML is missing.** Add `pyyaml` to `requirements.txt` at the repo "
             "root, commit, and let Streamlit rebuild.")
    st.stop()

HERE = Path(__file__).resolve().parent
CONTENT = HERE / "content"
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))          # so learner code (and the app) can `import sandbox`

# surface Streamlit-Cloud secrets to os.environ for the real-LLM modes
for _sec in ("OPENAI_API_KEY", "ANTHROPIC_API_KEY", "OPENAI_MODEL", "ANTHROPIC_MODEL",
             "FASTEMBED_MODEL"):
    try:
        if _sec in st.secrets and not os.environ.get(_sec):
            os.environ[_sec] = str(st.secrets[_sec])
    except Exception:
        pass

import sandbox  # noqa: E402  (needs the secret env vars + sys.path set above)

st.set_page_config(page_title="GenAI Lab Practice", page_icon="🧪", layout="wide",
                   initial_sidebar_state="expanded")

# canonical pip name for an import name (for the "missing package" message)
PIP_NAME = {
    "langchain_core": "langchain-core", "langchain_community": "langchain-community",
    "langchain_text_splitters": "langchain-text-splitters",
    "langchain_openai": "langchain-openai", "langchain_anthropic": "langchain-anthropic",
    "langgraph": "langgraph", "pypdf": "pypdf", "chromadb": "chromadb",
    "langchain_chroma": "langchain-chroma", "fastembed": "fastembed",
    "mcp": "mcp", "langchain_mcp_adapters": "langchain-mcp-adapters",
}

# mode -> (extra import needed to Run, env var holding the key)
MODE_REQUIRES = {
    "openai": ("langchain_openai", "OPENAI_API_KEY"),
    "claude": ("langchain_anthropic", "ANTHROPIC_API_KEY"),
}


# --------------------------------------------------------------------------- #
#  content                                                                     #
# --------------------------------------------------------------------------- #
@st.cache_data(show_spinner=False)
def list_labs() -> list[str]:
    if not CONTENT.is_dir():
        return []
    return sorted(p.name for p in CONTENT.iterdir() if (p / "practice.yaml").is_file())


@st.cache_data(show_spinner=False)
def load_lab(lab_id: str) -> dict:
    return yaml.safe_load((CONTENT / lab_id / "practice.yaml").read_text("utf-8"))


# --------------------------------------------------------------------------- #
#  grading (Check)                                                             #
# --------------------------------------------------------------------------- #
def code_only(src: str) -> str:
    out, in_str, q = [], False, ""
    for line in src.splitlines():
        buf, i = "", 0
        while i < len(line):
            chunk = line[i:i + 3]
            if not in_str and chunk in ('"""', "'''"):
                in_str, q, buf, i = True, chunk, buf + chunk, i + 3
                continue
            if in_str and line[i:i + 3] == q:
                in_str, buf, i = False, buf + q, i + 3
                continue
            if not in_str and line[i] == "#":
                break
            buf += line[i]
            i += 1
        out.append(buf)
    return "\n".join(out)


def check(code: str, checks: list[dict]) -> list[tuple[str, bool, int, int]]:
    clean = code_only(code)
    return [(c["desc"], bool(re.search(c["pattern"], clean, re.I | re.S)),
             c["points"] if re.search(c["pattern"], clean, re.I | re.S) else 0,
             c["points"]) for c in checks]


# each lab has 5 tasks; these weights sum to 100
TASK_WEIGHTS = {1: 10, 2: 25, 3: 25, 4: 15, 5: 25}


def task_weight(t: dict, idx: int) -> int:
    return int(t.get("weight", TASK_WEIGHTS.get(t.get("id", idx + 1), 20)))


def grade_task(code: str, t: dict, idx: int) -> dict:
    """Run the rubric, then scale points so the task totals `task_weight` and all
    tasks together total 100."""
    rows = check(code, t["checks"])                 # (desc, ok, pts_raw, max_raw)
    raw_max = sum(r[3] for r in rows) or 1
    w = task_weight(t, idx)
    scaled = [(d, ok, (m * w / raw_max if ok else 0.0), m * w / raw_max)
              for d, ok, _p, m in rows]
    return {"rows": scaled, "got": sum(r[2] for r in scaled), "max": float(w),
            "weight": w}


# --------------------------------------------------------------------------- #
#  Run - REAL execution in this interpreter                                    #
# --------------------------------------------------------------------------- #
def dependency_report(requires: list[str], mode: str) -> list[str]:
    need = list(requires or [])
    extra = MODE_REQUIRES.get(mode, (None, None))[0]
    if extra and extra not in need:
        need.append(extra)
    return [m for m in need if importlib.util.find_spec(m) is None]


PREAMBLE_IMPORTS = (
    "from sandbox import (get_chat_model, get_embeddings, get_tools, split_docs,\n"
    "                     load_sample_docs, make_vectorstore, supports_tool_calling,\n"
    "                     mcp_stdio_config, resolve_mode,\n"
    "                     SAMPLE_PDF, QUESTION_SET_PDF, MCP_SERVER_PATH)\n"
)


def preamble(mode: str) -> str:
    return (
        "import os as _os, sys as _sys\n"
        f"_sys.path.insert(0, r'{HERE}')\n"
        f"_os.environ['SANDBOX_MODE'] = '{mode}'\n"
        "import sandbox\n"
        + PREAMBLE_IMPORTS
        + f"MODE = '{mode}'\n"
        "# ---------------- your code below ----------------\n"
    )


def editor_key(task_id) -> str:
    return f"ed_{task_id}"


def keep_editors_alive(spec: dict) -> None:
    """Re-assign each task-editor's session_state value so Streamlit does NOT
    garbage-collect it when that task's widget isn't on screen. This is what
    makes typed code survive navigating to another task."""
    for t in spec["tasks"]:
        k = editor_key(t["id"])
        if k in st.session_state:
            st.session_state[k] = st.session_state[k]


def assemble_source(spec: dict, upto_id: int, mode: str) -> tuple[str, int]:
    parts = []
    for t in spec["tasks"]:
        parts.append(f"# ===== Task {t['id']} — {t['title']} =====\n"
                     + st.session_state.get(editor_key(t["id"]), t["starter"]))
        if t["id"] == upto_id:
            break
    return preamble(mode) + "\n\n".join(parts), len(parts)


def run_accumulated(spec: dict, upto_id: int, mode: str) -> dict:
    source, n_blocks = assemble_source(spec, upto_id, mode)
    demo = iter(["What is the standard notice period for regular full-time employees?",
                 "How many casual leave days are provided?", "exit", "exit", "exit"])
    ns = {"__name__": "__practice__", "input": lambda *_a: next(demo, "exit")}
    # a REAL temp file (has fileno) so subprocess-spawning code — the MCP lab —
    # can inherit stdio; StringIO would raise io.UnsupportedOperation: fileno
    cap = tempfile.TemporaryFile(mode="w+", encoding="utf-8", errors="replace")
    t0 = time.perf_counter()
    err = ""
    try:
        with contextlib.redirect_stdout(cap), contextlib.redirect_stderr(cap):
            exec(compile(source, "<practice>", "exec"), ns)  # noqa: S102
    except SystemExit:
        pass
    except BaseException:  # noqa: BLE001  — show EVERY real failure, never swallow
        err = traceback.format_exc()
    cap.seek(0)
    out = cap.read()
    cap.close()
    return {
        "out": out, "err": err, "source": source,
        "blocks": n_blocks, "seconds": time.perf_counter() - t0,
        "status": "error" if err else "ok",
    }


# --------------------------------------------------------------------------- #
#  UI                                                                          #
# --------------------------------------------------------------------------- #
def main() -> None:
    labs = list_labs()
    if not labs:
        st.error("No content found. Add `content/<lab>/practice.yaml`.")
        st.stop()

    qp = st.query_params.get("lab")
    with st.sidebar:
        if len(labs) > 1:
            lab_id = st.selectbox("Lab", labs, index=labs.index(qp) if qp in labs else 0)
        else:
            lab_id = labs[0]
        spec = load_lab(lab_id)
        keep_editors_alive(spec)   # <- typed code persists across task navigation
        st.title(spec.get("title", lab_id))

        st.divider()
        _MODE_LABELS = ["🔷 OpenAI (real)", "🟣 Claude (real)", "🧪 Mock (demo, no key)"]
        _MODE_VALS = {"🔷 OpenAI (real)": "openai", "🟣 Claude (real)": "claude",
                      "🧪 Mock (demo, no key)": "mock"}
        # default = REAL when a key exists, else Mock
        _auto = sandbox.resolve_mode("auto")
        _default_i = {"openai": 0, "claude": 1, "mock": 2}[_auto]
        pick = st.radio("LLM mode", _MODE_LABELS, index=_default_i,
                        help="Primary experience is REAL. Mock is a keyless demo "
                             "fallback — its LLM cannot do tool calling, so the "
                             "agent labs need OpenAI or Claude.")
        mode = _MODE_VALS[pick]
        if mode == "mock":
            st.caption("Demo fallback. **Real**: Python exec, LangChain, ChromaDB, "
                       "FastEmbed embeddings. **Stand-in**: only the answer LLM "
                       "(no tool calling).")
        else:
            pkg, envk = MODE_REQUIRES[mode]
            has_key = os.environ.get(envk, "").startswith(("sk-", "sk-ant-"))
            if importlib.util.find_spec(pkg) is None:
                st.caption(f"⚠ `{PIP_NAME[pkg]}` not installed — add it to "
                           f"`requirements.txt` and redeploy.")
            elif has_key:
                st.caption(f"✅ `{envk}` found — real "
                           f"{'OpenAI' if mode == 'openai' else 'Claude'} calls. "
                           + ("OpenAI embeddings too." if mode == "openai"
                              else "Embeddings use the local FastEmbed model."))
            else:
                st.caption(f"⚠ No `{envk}`. Set it as an env var or in "
                           f"**Manage app → Settings → Secrets**, or use Mock mode.")
        st.session_state["mode"] = mode

        if spec.get("notes"):
            st.divider()
            st.info(spec["notes"])

    # --- navigation in the MAIN area (works even when the sidebar is collapsed,
    #     e.g. on a narrow / mobile screen) ---
    kind, payload = nav_bar(spec)
    if kind == "overview":
        render_overview(spec)
    elif kind == "score":
        render_score(spec)
    else:
        render_task(spec, payload, mode)


def nav_bar(spec: dict):
    """Persistent Prev / selector / Next stepper shown above every page.
    Returns ("overview", None) | ("task", task_dict) | ("score", None)."""
    tasks = spec["tasks"]
    has_ov = bool(spec.get("overview"))
    options = (["📋 Overview"] if has_ov else []) \
        + [f"Task {t['id']} · {t['title']}" for t in tasks] + ["★ Score"]
    last = len(options) - 1
    i = max(0, min(last, st.session_state.setdefault("nav_i", 0)))

    c1, c2, c3 = st.columns([1, 6, 1])
    if c1.button("◀ Prev", use_container_width=True, disabled=i == 0):
        st.session_state["nav_i"] = i - 1
        st.rerun()
    if c3.button("Next ▶", use_container_width=True, disabled=i == last):
        st.session_state["nav_i"] = i + 1
        st.rerun()
    sel = c2.selectbox("Go to", options, index=i, label_visibility="collapsed")
    if options.index(sel) != i:
        st.session_state["nav_i"] = options.index(sel)
        st.rerun()
    st.progress(i / last if last else 0.0, text=f"{i + 1} / {len(options)}")

    if options[i] == "★ Score":
        return "score", None
    if options[i] == "📋 Overview":
        return "overview", None
    return "task", tasks[i - (1 if has_ov else 0)]


def render_overview(spec: dict) -> None:
    st.title(spec.get("title", "Overview"))
    st.markdown(spec.get("overview", "_No overview provided for this lab._"))
    if st.button("Start Task 1  ▶", type="primary"):
        st.session_state["nav_i"] = 1 if spec.get("overview") else 0
        st.rerun()


def band_for(pct: float) -> str:
    return ("Excellent" if pct >= 85 else "Pass" if pct >= 70
            else "Needs work" if pct >= 50 else "Incomplete")


def score_rows(spec: dict) -> tuple[list[dict], float]:
    """Per-task result rows (weighted to /100) + the total got."""
    subs = st.session_state.get("submissions", {})
    rows, got = [], 0.0
    for idx, t in enumerate(spec["tasks"]):
        tid = t["id"]
        w = task_weight(t, idx)
        sub = subs.get(tid)
        checked = st.session_state.get(f"result_{tid}")
        revealed = bool(st.session_state.get(f"revealed_{tid}", False))
        if sub:
            tgot, status, ts = sub["got"], "submitted", sub["ts"]
        elif checked:
            tgot, status, ts = sum(r[2] for r in checked), "checked (not submitted)", ""
        else:
            tgot, status, ts = 0.0, "not attempted", ""
        got += tgot
        rows.append({"id": tid, "title": t["title"], "weight": w,
                     "score": round(tgot, 1), "status": status,
                     "solution_viewed": "yes" if revealed else "no", "submitted_at": ts})
    return rows, got


def build_result_xlsx(spec: dict, rows: list[dict], total: float, pct: int, band: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    for k, v in (("Lab", spec.get("title", spec.get("lab_id"))),
                 ("Lab id", spec.get("lab_id")),
                 ("Generated", time.strftime("%Y-%m-%d %H:%M:%S")),
                 ("Total score", f"{round(total, 1)} / 100"),
                 ("Percent", f"{pct}%"), ("Band", band)):
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    header = ["Task", "Title", "Weight", "Score", "Status", "Solution viewed", "Submitted at"]
    ws.append(header)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["id"], r["title"], r["weight"], r["score"], r["status"],
                   r["solution_viewed"], r["submitted_at"]])
    ws.append([])
    ws.append(["", "TOTAL", sum(r["weight"] for r in rows), round(total, 1)])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    for col, width in zip("ABCDEFG", (8, 40, 8, 8, 22, 15, 20)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_score(spec: dict) -> None:
    st.title("Your score")
    rows, got = score_rows(spec)
    possible = sum(r["weight"] for r in rows)      # == 100
    n_submitted = sum(1 for r in rows if r["status"] == "submitted")

    for r in rows:
        mark = ("✔ " if r["status"] == "submitted"
                else "• " if r["status"].startswith("checked") else "· ")
        flag = "  — solution viewed" if r["solution_viewed"] == "yes" else ""
        tail = f"  ({r['submitted_at']})" if r["submitted_at"] else ""
        st.write(f"{mark}**Task {r['id']} — {r['title']}**  ·  "
                 f"**{r['score']} / {r['weight']}**  ·  {r['status']}{tail}{flag}")

    pct = round(100 * got / possible) if possible else 0
    band = band_for(pct)
    st.divider()
    st.subheader(f"{round(got, 1)} / 100   ({pct}%)   —   {band}")
    st.caption(f"Weights: T1 10 · T2 25 · T3 25 · T4 15 · T5 25   ·   "
               f"{n_submitted} / {len(rows)} tasks submitted")
    st.progress(pct / 100)

    try:
        xlsx = build_result_xlsx(spec, rows, got, pct, band)
        st.download_button(
            "⬇ Download result (Excel)", data=xlsx,
            file_name=f"{spec.get('lab_id', 'lab')}-result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary")
    except ModuleNotFoundError:
        st.warning("Excel export needs `openpyxl` — add it to requirements.txt. "
                   "Falling back to CSV.")
        csv = "Task,Title,Weight,Score,Status,Solution viewed,Submitted at\n" + "\n".join(
            f'{r["id"]},"{r["title"]}",{r["weight"]},{r["score"]},{r["status"]},'
            f'{r["solution_viewed"]},{r["submitted_at"]}' for r in rows
        ) + f'\n,,{"":s}TOTAL,{round(got, 1)}\n'
        st.download_button("⬇ Download result (CSV)", data=csv,
                           file_name=f"{spec.get('lab_id', 'lab')}-result.csv",
                           mime="text/csv")

    with st.expander("Also as JSON"):
        payload = {"lab": spec.get("lab_id"),
                   "generated": time.strftime("%Y-%m-%dT%H:%M:%S"),
                   "total": round(got, 1), "possible": 100, "percent": pct, "band": band,
                   "tasks": rows}
        st.code(json.dumps(payload, indent=2, ensure_ascii=False), language="json")


def render_task(spec: dict, t: dict, mode: str) -> None:
    tid = t["id"]
    hkey, rkey = f"hints_{tid}", f"revealed_{tid}"
    edk = editor_key(tid)
    # apply a deferred editor change (Reset / Copy) BEFORE the widget mounts —
    # you cannot assign a widget key's value after its widget renders this run
    pend = st.session_state.pop(f"_pending_{tid}", None)
    if pend is not None:
        st.session_state[edk] = pend
    st.session_state.setdefault(edk, t["starter"])
    st.session_state.setdefault(hkey, 0)
    subs = st.session_state.setdefault("submissions", {})

    idx = spec["tasks"].index(t)
    weight = task_weight(t, idx)
    st.title(f"Task {tid} — {t['title']}")
    st.caption(f"Weight: **{weight} marks** of 100")
    st.markdown(t["prompt"])
    if tid in subs:
        s = subs[tid]
        st.success(f"✔ Submitted **{s['got']:.1f} / {weight}** at {s['ts']} — "
                   "edit below and press **Submit** again to update it.")

    left, right = st.columns([3, 2])
    with left:
        st.markdown("**Your code**  ·  _kept as you move between tasks_")
        st.text_area("code", key=edk, height=340, label_visibility="collapsed")
        code = st.session_state[edk]

        r1c1, r1c2, r1c3 = st.columns(3)
        do_check = r1c1.button("✓ Check", key=f"chk_{tid}",
                               help="Score this task's code against the rubric.")
        do_submit = r1c2.button("✔ Submit", type="primary", key=f"sub_{tid}",
                                help="Save this answer + its score. You can resubmit.")
        do_run = r1c3.button(f"▶ Run ({mode})", key=f"run_{tid}",
                             help="Execute Task 1..N in this app's Python interpreter.")
        r2c1, r2c2, _ = st.columns(3)
        if r2c1.button("💡 Hint", key=f"hint_{tid}"):
            st.session_state[hkey] = min(len(t["hints"]), st.session_state[hkey] + 1)
        if r2c2.button("↺ Reset", key=f"rst_{tid}"):
            st.session_state[f"_pending_{tid}"] = t["starter"]
            st.session_state.pop(f"result_{tid}", None)
            st.rerun()

        if do_check or do_submit:
            st.session_state[f"result_{tid}"] = grade_task(code, t, idx)["rows"]
        if do_submit:
            rows = st.session_state[f"result_{tid}"]
            subs[tid] = {"code": code,
                         "rows": [(d, ok, round(p, 1), round(m, 1)) for d, ok, p, m in rows],
                         "got": round(sum(r[2] for r in rows), 1), "max": weight,
                         "revealed": bool(st.session_state.get(rkey, False)),
                         "ts": time.strftime("%Y-%m-%d %H:%M:%S")}
            st.toast(f"Task {tid} submitted — {subs[tid]['got']:.1f} / {weight}", icon="✔")
            st.rerun()

        rows = st.session_state.get(f"result_{tid}")
        if rows:
            got = sum(r[2] for r in rows)
            st.markdown(f"**Check: {got:.1f} / {weight} marks**")
            for desc, ok, pts, pmax in rows:
                st.markdown(f"{'✅' if ok else '❌'} {desc} — {pts:.1f} / {pmax:.1f}")

        needs_real = (spec.get("mode_required") == "real") or t.get("needs_real")
        if needs_real and mode == "mock":
            st.info("**This exercise needs a real tool-calling LLM.** Switch to "
                    "**OpenAI** or **Claude** mode in the sidebar (set "
                    "`OPENAI_API_KEY` / `ANTHROPIC_API_KEY` first). Mock's model "
                    "cannot emit tool calls.")

        if do_run:
            missing = dependency_report(spec.get("requires", []), mode)
            envk = MODE_REQUIRES.get(mode, (None, None))[1]
            has_key = bool(envk) and os.environ.get(envk, "").startswith(("sk-", "sk-ant-"))
            if needs_real and mode == "mock":
                st.error("Cannot Run in Mock mode — this exercise requires OpenAI or "
                         "Claude mode. Nothing was executed or simulated.")
            elif missing:
                st.error(
                    "**Cannot Run — packages missing from this app's Python "
                    "environment:**\n\n"
                    + "\n".join(f"- `{PIP_NAME.get(m, m)}` (import `{m}`)" for m in missing)
                    + "\n\nAdd them to `requirements.txt` and rebuild (cloud) or "
                    "`pip install -r requirements.txt` (local). Nothing is simulated.")
            elif envk and not has_key:
                st.error(f"**{mode.title()} mode needs `{envk}`.** Set it as an env "
                         f"var / Streamlit secret, or switch to Mock mode.")
            else:
                with st.spinner(f"Running your code — real execution ({mode})…"):
                    res = run_accumulated(spec, tid, mode)

                icon = "✅" if res["status"] == "ok" else "❌"
                st.markdown(f"{icon} **{res['status'].upper()}** · `{mode}` mode · "
                            f"{res['blocks']} task block(s) · {res['seconds']:.2f}s")

                with st.expander("Execution steps — the exact code that ran", expanded=False):
                    st.caption("preamble (injected) + Task 1…N, concatenated and `exec`'d "
                               "in this app's Python interpreter")
                    st.code(res["source"], language="python")

                st.markdown("**Actual output** (real stdout / stderr)")
                st.code(res["out"] or "(no output)", language="text")
                if res["err"]:
                    st.markdown("**Error** — real traceback from your code, not simulated")
                    st.code(res["err"], language="text")

    with right:
        shown = st.session_state[hkey]
        st.markdown(f"**Hints** ({shown}/{len(t['hints'])})")
        for h in t["hints"][:shown]:
            st.markdown(f"- {h}")
        if shown < len(t["hints"]):
            st.caption("Press **💡 Hint** for the next one.")
        st.divider()
        if st.checkbox("Reveal solution code", key=f"revchk_{tid}",
                       value=st.session_state.get(rkey, False)):
            st.session_state[rkey] = True
            st.code(t["solution"], language="python")
            if st.button("Copy into my editor", key=f"cpy_{tid}"):
                st.session_state[f"_pending_{tid}"] = t["solution"]
                st.rerun()
        else:
            st.caption("Try it yourself first — revealing is flagged on the Score page.")


if __name__ == "__main__":
    main()
